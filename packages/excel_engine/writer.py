"""Excel report writer using openpyxl."""

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass
class ReportSheet:
    """Definition of a sheet to write to an Excel report."""

    name: str
    headers: list[str]
    rows: list[list[Any]] = field(default_factory=list)
    column_widths: dict[int, int] = field(default_factory=dict)


class ExcelWriter:
    """Writer for generating formatted Excel reports."""

    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    def write_report(
        self,
        output_path: Path,
        sheets: list[ReportSheet],
    ) -> Path:
        """Write a multi-sheet Excel report.

        Args:
            output_path: Destination file path.
            sheets: List of sheet definitions.

        Returns:
            Path to the written file.

        Raises:
            ValueError: If writing fails.
        """
        logger.info("Writing Excel report to: {}", output_path)

        try:
            workbook = Workbook()
            default_sheet = workbook.active
            if default_sheet is not None:
                workbook.remove(default_sheet)

            for sheet_def in sheets:
                self._write_sheet(workbook, sheet_def)

            output_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(output_path)
            logger.info("Excel report written: {} sheets", len(sheets))
            return output_path

        except Exception as exc:
            logger.error("Failed to write Excel report: {}", exc)
            raise ValueError(f"Excel write failed: {exc}") from exc

    def _write_sheet(self, workbook: Workbook, sheet_def: ReportSheet) -> None:
        """Write a single sheet to the workbook.

        Args:
            workbook: openpyxl Workbook instance.
            sheet_def: Sheet definition with headers and rows.
        """
        sheet = workbook.create_sheet(title=sheet_def.name)

        for col_idx, header in enumerate(sheet_def.headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT

        for row_idx, row_data in enumerate(sheet_def.rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        for col_idx, width in sheet_def.column_widths.items():
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

        sheet.freeze_panes = "A2"
