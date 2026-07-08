"""Excel file reader using openpyxl and Polars."""

from pathlib import Path
from typing import Any, cast

import polars as pl
from loguru import logger
from openpyxl import load_workbook


class ExcelReader:
    """Reader for Excel workbook files (.xlsx, .xls)."""

    SUPPORTED_EXTENSIONS: tuple[str, ...] = (".xlsx", ".xls")

    def read_sheet(
        self,
        file_path: Path,
        sheet_name: str | None = None,
    ) -> pl.DataFrame:
        """Read a specific sheet from an Excel workbook.

        Args:
            file_path: Path to the Excel file.
            sheet_name: Optional sheet name (defaults to first sheet).

        Returns:
            Polars DataFrame with sheet contents.

        Raises:
            ValueError: If file cannot be read.
        """
        logger.info("Reading Excel sheet from: {}", file_path.name)

        try:
            if sheet_name is not None:
                return pl.read_excel(file_path, sheet_name=sheet_name)
            return cast("pl.DataFrame", pl.read_excel(file_path, sheet_id=0))
        except Exception as exc:
            logger.error("Failed to read Excel with Polars: {}", exc)
            return self._read_with_openpyxl(file_path, sheet_name)

    def _read_with_openpyxl(
        self,
        file_path: Path,
        sheet_name: str | None,
    ) -> pl.DataFrame:
        """Fallback Excel reader using openpyxl."""
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            target_sheet = sheet_name or workbook.sheetnames[0]

            if target_sheet not in workbook.sheetnames:
                raise ValueError(f"Sheet not found: {target_sheet}")

            sheet = workbook[target_sheet]
            rows: list[list[Any]] = []
            for row in sheet.iter_rows(values_only=True):
                rows.append(list(row))

            if not rows:
                return pl.DataFrame()

            headers = [
                str(header) if header is not None else f"col_{index}"
                for index, header in enumerate(rows[0])
            ]
            data_rows = rows[1:]
            return pl.DataFrame(
                {
                    headers[index]: [row[index] for row in data_rows]
                    for index in range(len(headers))
                },
            )

        except Exception as exc:
            logger.error("openpyxl read failed: {}", exc)
            raise ValueError(f"Failed to read Excel file: {exc}") from exc

    def list_sheets(self, file_path: Path) -> list[str]:
        """List all sheet names in an Excel workbook."""
        try:
            workbook = load_workbook(file_path, read_only=True)
            return cast("list[str]", list(workbook.sheetnames))
        except Exception as exc:
            logger.error("Failed to list sheets: {}", exc)
            raise ValueError(f"Failed to read workbook: {exc}") from exc
